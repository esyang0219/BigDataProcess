#!/usr/bin/python3
from openpyxl import load_workbook
wb = load_workbook(filename='student.xlsx')
ws = wb.active

rank=[]
list=[]
s_list=[]
f_cnt=0

for row in range(2, ws.max_row+1):
	mid = ws.cell(row = row, column = 3).value
	fin = ws.cell(row = row, column = 4).value
	hw = ws.cell(row = row, column = 5).value
	at = ws.cell(row = row, column = 6).value
	total = mid*0.3+fin*0.35+hw*0.34+at
	ws.cell(row = row, column = 7, value = total)
	list.append(ws.cell(row=row, column=7).value)
s_list=sorted(list)
s_list.reverse()
for row in list:
	rank.append(s_list.index(row)+1)
n = ws.max_row-1
a_value=int(n*0.3)
b_value=int(n*0.7)
for row in range(2, ws.max_row+1):
	if ws.cell(row=row, column=7).value < 40:
		f_cnt += 1
cp_value=int(n*0.7+((0.3*n-f_cnt)/2))

for row in range(2, ws.max_row+1):
	if ws.cell(row=row, column=7).value >= 40:
		if rank[row-2] <= (a_value / 2):
			grade='A+'
		elif rank[row-2] <= a_value:
			grade='A0'
		elif rank[row-2] <= (b_value / 2):
			grade='B+'
		elif rank[row-2] <= b_value:
			grade='B0'
		elif rank[row-2] <= cp_value:
			grade='C+'
		else:
			grade='C0'			
	else:
		grade='F'	
	ws.cell(row=row, column=8, value=grade)
wb.save('student.xlsx')
